"""測試新加的：寫回 Excel、鎖檔偵測、設定遷移、盤中時段判斷。"""
import datetime as dt
import shutil
import sys
import tempfile
from pathlib import Path
from zoneinfo import ZoneInfo

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook, load_workbook

from core import config as C

TMP = Path(tempfile.mkdtemp(prefix="finance_edit_"))
C.HOLDINGS_XLSX = TMP / "持股.xlsx"
C.DATA_DIR = TMP / "data"
C.SETTINGS_JSON = C.DATA_DIR / "設定.json"
C.HISTORY_XLSX = C.DATA_DIR / "歷史紀錄.xlsx"
C.EXPORT_XLSX = C.DATA_DIR / "庫存總覽.xlsx"
C.BACKUP_DIR = C.DATA_DIR / "備份"
C.CACHE_JSON = C.DATA_DIR / "報價快取.json"

from core import excel_io, market_hours                      # noqa: E402
from core import settings as settings_store                  # noqa: E402
from core.excel_io import HoldingsFileError                  # noqa: E402
import app as app_module                                     # noqa: E402

FAIL = 0
TPE = ZoneInfo("Asia/Taipei")


def ok(name, cond, extra=""):
    global FAIL
    if cond:
        print(f"  PASS  {name}" + (f"  {extra}" if extra else ""))
    else:
        FAIL += 1
        print(f"  FAIL  {name}  {extra}")


def reset():
    for p in (C.HOLDINGS_XLSX, C.SETTINGS_JSON):
        if p.exists():
            p.unlink()


# ══════════════════════════════════════════════════════════════════
print("=" * 74)
print("1. 寫回 持股.xlsx")
print("=" * 74)
reset()
excel_io.ensure_holdings_template()

rows = [
    {C.COL_CODE: "2882", C.COL_NAME: "國泰金", C.COL_MARKET: "TW",
     C.COL_SHARES: 217, C.COL_AVGCOST: 44.66, C.COL_NOTE: "測試"},
    {C.COL_CODE: "00919", C.COL_NAME: "", C.COL_MARKET: "TW",
     C.COL_SHARES: 38000, C.COL_AVGCOST: 22.24, C.COL_NOTE: ""},
    {C.COL_CODE: "AAPL", C.COL_NAME: "Apple", C.COL_MARKET: "US",
     C.COL_SHARES: 50, C.COL_AVGCOST: 180.5, C.COL_NOTE: ""},
]
info = excel_io.write_holdings(rows)
ok("寫回成功", info["count"] == 3)

book = excel_io.read_holdings()
ok("讀回來還是 3 檔", len(book.holdings) == 3)
by = {h.code: h for h in book.holdings}
ok("00919 沒有被 Excel 變成 919", "00919" in by)
ok("股數正確", by["00919"].shares == 38000.0)
ok("均價正確", abs(by["2882"].avg_cost - 44.66) < 1e-9)
ok("備註保留", by["2882"].note == "測試")
ok("美股市場別保留", by["AAPL"].market == C.MARKET_US)

wb = load_workbook(C.HOLDINGS_XLSX)
ws = wb[C.SHEET_HOLDINGS]
ok("表頭正確", [c.value for c in ws[1]] == C.HOLDING_COLUMNS)
data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value not in (None, "")]
ok("沒有多餘的資料列", data_rows == [2, 3, 4], f"有資料的列={data_rows}")
ok("預留的空白列沒有開太多", ws.max_row <= 4 + 25, f"max_row={ws.max_row}")
ok("代號欄是文字格式", ws.cell(row=2, column=1).number_format == "@")
ok("凍結窗格保留", ws.freeze_panes == "A2")
wb.close()

ok("寫入前有備份", C.BACKUP_DIR.exists()
   and any(p.name.startswith("持股_") for p in C.BACKUP_DIR.iterdir()))
ok("沒有殘留 .tmp", not list(TMP.rglob("*.tmp")))

print("\n  改成 2 檔（刪掉一檔）:")
excel_io.write_holdings(rows[:2])
ok("刪除的那列真的不見了", len(excel_io.read_holdings().holdings) == 2)


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 74)
print("2. 外部改動偵測（避免蓋掉別人的修改）")
print("=" * 74)
book = excel_io.read_holdings()
old_mtime = book.mtime

import time                                                   # noqa: E402
time.sleep(1.1)
excel_io.write_holdings(rows)                                 # 模擬別的地方改了檔案

try:
    excel_io.write_holdings(rows[:1], expect_mtime=old_mtime)
    ok("用過期的 mtime 存檔應該被擋", False)
except HoldingsFileError as e:
    ok("用過期的 mtime 存檔被擋下", "被其他程式改過" in str(e))

fresh = excel_io.read_holdings().mtime
excel_io.write_holdings(rows[:1], expect_mtime=fresh)
ok("用最新的 mtime 就存得進去", len(excel_io.read_holdings().holdings) == 1)


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 74)
print("3. Excel 鎖檔偵測")
print("=" * 74)
ok("沒開 Excel 時偵測不到鎖檔", excel_io.excel_lock_holder(C.HOLDINGS_XLSX) is None)
lock = C.HOLDINGS_XLSX.with_name("~$" + C.HOLDINGS_XLSX.name)
lock.write_text("x", encoding="utf-8")
ok("有 ~$ 檔時偵測得到", excel_io.excel_lock_holder(C.HOLDINGS_XLSX) is not None)
lock.unlink()


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 74)
print("4. 存檔驗證（透過 Api，跟前端走同一條路）")
print("=" * 74)
api = app_module.Api()

cases = [
    ("代號空白", [{C.COL_CODE: "", C.COL_SHARES: 100, C.COL_AVGCOST: 10}], C.COL_CODE, "不能空白"),
    ("股數是文字", [{C.COL_CODE: "2882", C.COL_SHARES: "abc", C.COL_AVGCOST: 10}], C.COL_SHARES, "大於 0"),
    ("股數是 0", [{C.COL_CODE: "2882", C.COL_SHARES: 0, C.COL_AVGCOST: 10}], C.COL_SHARES, "大於 0"),
    ("均價是負數", [{C.COL_CODE: "2882", C.COL_SHARES: 10, C.COL_AVGCOST: -5}], C.COL_AVGCOST, "0 或正數"),
]
for name, bad, field, expect in cases:
    res = api.save_holdings(bad)
    err = (res.get("errors") or [{}])[0]
    ok(name, not res.get("ok") and err.get("field") == field and expect in err.get("msg", ""),
       f'-> {err.get("msg", res.get("error"))}')

print("\n  代號重複不是錯誤，而是先要求確認（詳細測試在 test_duplicates.py）:")
dup = [{C.COL_CODE: "2882", C.COL_SHARES: 10, C.COL_AVGCOST: 10},
       {C.COL_CODE: "2882", C.COL_SHARES: 20, C.COL_AVGCOST: 20}]
res = api.save_holdings(dup)
ok("回 needs_confirm 而不是 errors",
   res.get("needs_confirm") == "duplicates" and not res.get("errors"), str(res)[:80])
ok("確認後存得進去", api.save_holdings(dup, confirmed=True).get("ok"))

print("\n  容錯與自動判斷:")
res = api.save_holdings([
    {C.COL_CODE: " 2882 ", C.COL_SHARES: "38,000", C.COL_AVGCOST: "$44.66", C.COL_MARKET: ""},
    {C.COL_CODE: "AAPL", C.COL_SHARES: 50, C.COL_AVGCOST: 180.5, C.COL_MARKET: ""},
])
ok("千分位與 $ 能吃、市場自動判斷", res.get("ok"), str(res))
saved = {h.code: h for h in excel_io.read_holdings().holdings}
ok("代號空白被去掉", "2882" in saved)
ok("38,000 轉成數字", saved["2882"].shares == 38000.0)
ok("台股自動判為 TW", saved["2882"].market == C.MARKET_TW)
ok("美股自動判為 US", saved["AAPL"].market == C.MARKET_US)


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 74)
print("5. 設定：讀寫與遷移")
print("=" * 74)
reset()
s = settings_store.load()
ok("沒有設定檔時用預設值", s[C.SET_MANUAL_FX] is None and s[C.SET_AUTO_REFRESH] is True)

s = settings_store.save({C.SET_MANUAL_FX: "32.5", C.SET_COLOR_SCHEME: C.COLOR_US_STYLE,
                         C.SET_AUTO_REFRESH: False})
ok("手動匯率存成數字", s[C.SET_MANUAL_FX] == 32.5)
ok("顏色存得下", s[C.SET_COLOR_SCHEME] == C.COLOR_US_STYLE)
ok("自動更新關閉", s[C.SET_AUTO_REFRESH] is False)
ok("重讀一致", settings_store.load() == s)

ok("匯率填文字退回自動", settings_store.save({C.SET_MANUAL_FX: "abc"})[C.SET_MANUAL_FX] is None)
ok("匯率填 0 退回自動", settings_store.save({C.SET_MANUAL_FX: 0})[C.SET_MANUAL_FX] is None)
ok("顏色填怪東西退回台股", settings_store.save({C.SET_COLOR_SCHEME: "藍漲"})[C.SET_COLOR_SCHEME]
   == C.COLOR_TW_STYLE)

C.SETTINGS_JSON.write_text("{壞掉的 json", encoding="utf-8")
ok("設定檔壞掉不會讓程式掛掉", settings_store.load() == dict(C.DEFAULT_SETTINGS))

print("\n  從舊版 Excel 設定分頁遷移:")
reset()
wb = Workbook()
ws = wb.active
ws.title = C.SHEET_HOLDINGS
ws.append(C.HOLDING_COLUMNS)
ws.append(["2882", "", "TW", 100, 40, ""])
st = wb.create_sheet(C.SHEET_SETTINGS_LEGACY)
st.append(["設定項目", "值", "說明"])
st.append([C.SET_MANUAL_FX, 33.25, ""])
st.append([C.SET_COLOR_SCHEME, C.COLOR_US_STYLE, ""])
st.append([C.LEGACY_AUTO_REFRESH, 60, ""])
wb.save(C.HOLDINGS_XLSX)

ok("有搬", settings_store.migrate_from_excel() is True)
s = settings_store.load()
ok("手動匯率搬過來", s[C.SET_MANUAL_FX] == 33.25)
ok("顏色搬過來", s[C.SET_COLOR_SCHEME] == C.COLOR_US_STYLE)
ok("舊的 60 秒轉成「開啟」", s[C.SET_AUTO_REFRESH] is True)

wb = load_workbook(C.HOLDINGS_XLSX)
ok("Excel 的設定分頁已移除", C.SHEET_SETTINGS_LEGACY not in wb.sheetnames, str(wb.sheetnames))
ok("持股分頁還在且資料完整", C.SHEET_HOLDINGS in wb.sheetnames)
wb.close()
ok("持股讀得回來", len(excel_io.read_holdings().holdings) == 1)
ok("第二次不會重複搬", settings_store.migrate_from_excel() is False)


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 74)
print("6. 盤中時段判斷")
print("=" * 74)


def at(y, m, d, hh, mm):
    return dt.datetime(y, m, d, hh, mm, tzinfo=TPE)


# 2026-08-14 是星期五，2026-08-15 是星期六
ok("台股 週五 10:30 開盤中", market_hours.tw_open(at(2026, 8, 14, 10, 30)) is True)
ok("台股 週五 08:59 還沒開", market_hours.tw_open(at(2026, 8, 14, 8, 59)) is False)
ok("台股 週五 13:30 剛好收盤", market_hours.tw_open(at(2026, 8, 14, 13, 30)) is True)
ok("台股 週五 13:31 已收盤", market_hours.tw_open(at(2026, 8, 14, 13, 31)) is False)
ok("台股 週六 10:30 不開", market_hours.tw_open(at(2026, 8, 15, 10, 30)) is False)

# 美股夏令時間 EDT = 台北時間 21:30 - 04:00
ok("美股 台北時間週五 22:00 開盤中(夏令)", market_hours.us_open(at(2026, 8, 14, 22, 0)) is True)
ok("美股 台北時間週五 20:00 還沒開", market_hours.us_open(at(2026, 8, 14, 20, 0)) is False)
# 冬令時間 EST = 台北時間 22:30 - 05:00；2026-01-15 是星期四
ok("美股 台北時間 22:00 冬令還沒開", market_hours.us_open(at(2026, 1, 15, 22, 0)) is False)
ok("美股 台北時間 23:00 冬令開盤中", market_hours.us_open(at(2026, 1, 15, 23, 0)) is True)

print()
sec, why = market_hours.next_interval(True, False, now=at(2026, 8, 14, 10, 0))
ok("只有台股 · 盤中 -> 60 秒", sec == 60 and "台股盤中" in why, why)
sec, why = market_hours.next_interval(True, False, now=at(2026, 8, 14, 16, 0))
ok("只有台股 · 收盤後 -> 600 秒", sec == 600, f"{sec} {why}")
sec, why = market_hours.next_interval(True, True, now=at(2026, 8, 14, 22, 0))
ok("有美股 · 台北深夜 -> 60 秒", sec == 60 and "美股盤中" in why, why)
sec, why = market_hours.next_interval(True, False, now=at(2026, 8, 15, 10, 0))
ok("週六 -> 600 秒", sec == 600, f"{sec} {why}")

sec, why = market_hours.next_interval(True, False, tw_data_is_today=False,
                                      now=at(2026, 8, 14, 10, 0))
ok("時間在盤中但資料不是今天（休市）-> 600 秒", sec == 600 and "休市" in why, why)
sec, why = market_hours.next_interval(True, False, tw_data_is_today=True,
                                      now=at(2026, 8, 14, 10, 0))
ok("資料是今天 -> 維持 60 秒", sec == 60, f"{sec} {why}")
sec, _ = market_hours.next_interval(True, False, tw_data_is_today=None,
                                    now=at(2026, 8, 14, 10, 0))
ok("無從判斷時照時間走", sec == 60)


shutil.rmtree(TMP, ignore_errors=True)
print("\n" + "=" * 74)
print("全部通過" if not FAIL else f"有 {FAIL} 項失敗")
print("=" * 74)
sys.exit(1 if FAIL else 0)
