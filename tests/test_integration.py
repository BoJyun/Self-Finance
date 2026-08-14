"""整合測試：真的打 API、真的讀寫 Excel，把完整流程跑一遍。

用暫存目錄，不會動到使用者的 持股.xlsx。
"""
import sys, tempfile, shutil, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook, load_workbook

from core import config as C

# 把所有路徑改指向暫存目錄
TMP = Path(tempfile.mkdtemp(prefix="finance_test_"))
C.HOLDINGS_XLSX = TMP / "持股.xlsx"
C.DATA_DIR = TMP / "data"
C.HISTORY_XLSX = C.DATA_DIR / "歷史紀錄.xlsx"
C.EXPORT_XLSX = C.DATA_DIR / "庫存總覽.xlsx"
C.BACKUP_DIR = C.DATA_DIR / "備份"
C.CACHE_JSON = C.DATA_DIR / "報價快取.json"
C.SETTINGS_JSON = C.DATA_DIR / "設定.json"

from core import excel_io                                    # noqa: E402
from core import settings as settings_store                  # noqa: E402
from core.excel_io import HoldingsFileError                  # noqa: E402
import app as app_module                                     # noqa: E402

FAIL = 0


def ok(name, cond, extra=""):
    global FAIL
    if cond:
        print(f"  PASS  {name}" + (f"  {extra}" if extra else ""))
    else:
        FAIL += 1
        print(f"  FAIL  {name}  {extra}")


def make_holdings(rows, settings=None, header=None):
    wb = Workbook()
    ws = wb.active
    ws.title = C.SHEET_HOLDINGS
    ws.append(header or C.HOLDING_COLUMNS)
    for r in rows:
        ws.append(r)
    wb.save(C.HOLDINGS_XLSX)
    settings_store.save(settings or dict(C.DEFAULT_SETTINGS))


# ══════════════════════════════════════════════════════════════════
print("=" * 78)
print("1. 範本自動產生")
print("=" * 78)
created = excel_io.ensure_holdings_template()
ok("首次執行建立範本", created and C.HOLDINGS_XLSX.exists())
ok("第二次執行不覆蓋", excel_io.ensure_holdings_template() is False)
book = excel_io.read_holdings()
ok("範本讀得回來", len(book.holdings) == 2, f"{len(book.holdings)} 檔")
ok("範本的美股列判斷正確",
   any(h.code == "AAPL" and h.market == C.MARKET_US for h in book.holdings))
ok("範本不再產生設定分頁（設定已搬到 JSON）",
   C.SHEET_SETTINGS_LEGACY not in load_workbook(C.HOLDINGS_XLSX).sheetnames)
ok("設定預設值：手動匯率為空", settings_store.load()[C.SET_MANUAL_FX] is None)
C.HOLDINGS_XLSX.unlink()


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 78)
print("2. 錯誤處理：訊息要看得懂")
print("=" * 78)

cases = [
    ("缺少必要欄位", [["2882", "", "TW", 217, 44.66, ""]], ["代號", "名稱"], "缺少必要欄位"),
    ("股數不是數字", [["2882", "", "TW", "很多", 44.66, ""]], None, "不是有效的數字"),
    ("代號重複", [["2882", "", "TW", 100, 40, ""], ["2882", "", "TW", 200, 50, ""]], None, "出現了兩次"),
    ("股數是 0", [["2882", "", "TW", 0, 44.66, ""]], None, "必須大於 0"),
    ("均價空白", [["2882", "", "TW", 217, None, ""]], None, "都要填"),
    ("市場填錯", [["2882", "", "火星", 217, 44.66, ""]], None, "請填 TW"),
]
for name, rows, header, expect in cases:
    make_holdings(rows, header=header)
    try:
        excel_io.read_holdings()
        ok(name, False, "應該要報錯但沒有")
    except HoldingsFileError as e:
        ok(name, expect in str(e), f'-> "{str(e).splitlines()[0][:52]}…"')

print("\n  容錯（這些不該報錯）:")
make_holdings([
    ["2882", "", "TW", "38,000", " 44.66 ", ""],     # 千分位、空白
    [None, None, None, None, None, None],             # 空白列
    ["  00919  ", "群益", "台股", 100, "$30.57", "中文市場別"],
    ["AAPL", "", "", 50, 180.5, "市場留空自動判斷"],
])
book = excel_io.read_holdings()
ok("千分位/空白/空列/中文市場別/自動判斷都能處理", len(book.holdings) == 3, f"{len(book.holdings)} 檔")
by = {h.code: h for h in book.holdings}
ok("千分位轉數字", by["2882"].shares == 38000.0)
ok("代號前後空白被去掉", "00919" in by)
ok("「台股」認得出來", by["00919"].market == C.MARKET_TW)
ok("$30.57 轉數字", by["00919"].avg_cost == 30.57)
ok("AAPL 自動判為美股", by["AAPL"].market == C.MARKET_US)

make_holdings([["2882", "", "TW", 100, 40, ""]], settings={C.SET_MANUAL_FX: "abc"})
ok("手動匯率填錯會退回自動", settings_store.load()[C.SET_MANUAL_FX] is None)


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 78)
print("3. 完整流程：真的抓報價（台股 5 檔 + 美股 2 檔）")
print("=" * 78)
make_holdings([
    ["2882", "", "TW", 217, 44.66, ""],
    ["00919", "", "TW", 38000, 22.24, ""],
    ["00878", "", "TW", 46000, 23.92, ""],
    ["006208", "", "TW", 308, 91.97, ""],
    ["2885", "", "TW", 3518, 27.06, ""],
    ["AAPL", "", "US", 50, 180.5, ""],
    ["VOO", "", "US", 20, 500.0, ""],
])

api = app_module.Api()
payload = api.refresh()

if payload.get("error"):
    print("  ERROR:", payload["error"])
    FAIL += 1
else:
    s = payload["summary"]
    print(f"\n  {'代號':<9}{'名稱':<20}{'現價':>10}{'漲跌%':>9}{'今日損益':>12}"
          f"{'總損益':>13}{'報酬率':>9}{'市值(TWD)':>14}  來源")
    for r in payload["rows"]:
        f = lambda v, d=2: ("—" if v is None else f"{v:,.{d}f}")
        print(f"  {r['代號']:<9}{(r['名稱'] or '')[:18]:<20}{f(r['現價']):>10}"
              f"{f(r['漲跌幅']):>9}{f(r['今日損益'],0):>12}{f(r['總損益'],0):>13}"
              f"{f(r['報酬率']):>9}{f(r['市值'],0):>14}  {r['資料來源']}")

    print(f"\n  總市值 {s['market_value']:,.0f}   成本 {s['cost']:,.0f}")
    print(f"  今日損益 {s['today_pnl']:+,.0f} ({s['today_return']:+.2f}%)")
    print(f"  累積損益 {s['total_pnl']:+,.0f} ({s['total_return']:+.2f}%)")
    print(f"  台股 {s['tw_value']:,.0f} / 美股 {s['us_value']:,.0f} (佔 {s['us_ratio']:.1f}%)")
    print(f"  USD/TWD {s['fx_rate']} 來自 {s['fx_source']}")

    print()
    ok("7 檔都有回來", len(payload["rows"]) == 7)
    ok("每一檔都抓到現價", all(r["現價"] is not None for r in payload["rows"]),
       f"缺: {[r['代號'] for r in payload['rows'] if r['現價'] is None]}")
    ok("每一檔都抓到昨收", all(r["昨收"] is not None for r in payload["rows"]))
    ok("台股用證交所 MIS",
       all(r["資料來源"].startswith("證交所") for r in payload["rows"] if r["市場"] == "TW"))
    ok("美股用 Yahoo",
       all("Yahoo" in r["資料來源"] or "yfinance" in r["資料來源"]
           for r in payload["rows"] if r["市場"] == "US"))
    ok("抓到匯率", s["fx_rate"] and 20 < s["fx_rate"] < 45, f"{s['fx_rate']}")
    ok("沒有算不完整的標的", not s["incomplete"], str(s["incomplete"]))
    ok("名稱自動補上", all(r["名稱"] and r["名稱"] != r["代號"] for r in payload["rows"]))

    # 交叉驗算：總市值應等於各列市值之和
    manual = sum(r["市值"] for r in payload["rows"])
    ok("總市值 = 各列市值加總", abs(manual - s["market_value"]) < 0.01)
    tw = sum(r["股數"] * r["現價"] for r in payload["rows"] if r["市場"] == "TW")
    ok("台股市值手算對得上", abs(tw - s["tw_value"]) < 0.01, f"{tw:,.0f}")
    us = sum(r["股數"] * r["現價"] for r in payload["rows"] if r["市場"] == "US") * s["fx_rate"]
    ok("美股市值換算手算對得上", abs(us - s["us_value"]) < 0.01, f"{us:,.0f}")


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 78)
print("4. 寫入：歷史淨值、匯出、備份、原子替換")
print("=" * 78)
ok("每日淨值有寫入", C.HISTORY_XLSX.exists() and payload.get("nav_saved"))
wb = load_workbook(C.HISTORY_XLSX)
ws = wb[C.SHEET_NAV]
ok("淨值表頭正確", [c.value for c in ws[1]] == C.NAV_COLUMNS)
ok("淨值有一筆資料", ws.max_row == 2, f"max_row={ws.max_row}")
ok("日期是今天", str(ws.cell(row=2, column=1).value) == datetime.date.today().strftime("%Y-%m-%d"))
wb.close()

api.refresh()                                    # 同一天再跑一次
wb = load_workbook(C.HISTORY_XLSX)
ok("同一天重跑是覆蓋不是新增", wb[C.SHEET_NAV].max_row == 2, f"max_row={wb[C.SHEET_NAV].max_row}")
wb.close()
ok("有產生備份", C.BACKUP_DIR.exists() and any(C.BACKUP_DIR.iterdir()))

res = api.export_overview(payload["rows"], payload["summary"])
ok("庫存總覽匯出成功", res.get("ok") and C.EXPORT_XLSX.exists())
if C.EXPORT_XLSX.exists():
    wb = load_workbook(C.EXPORT_XLSX)
    ws = wb.active
    ok("匯出檔有 7 檔 + 表頭 + 總計", ws.max_row >= 11, f"max_row={ws.max_row}")
    wb.close()

ok("沒有殘留 .tmp 檔", not list(C.DATA_DIR.rglob("*.tmp")))
ok("報價快取有寫出", C.CACHE_JSON.exists())
ok("快取讀得回來", (excel_io.load_cache() or {}).get("summary") is not None)

ok("持股.xlsx 只有「持股」分頁", load_workbook(C.HOLDINGS_XLSX).sheetnames == [C.SHEET_HOLDINGS])

# 更新報價這條路徑絕不能寫使用者的檔案，否則 Excel 開著時每次更新都會失敗
before = C.HOLDINGS_XLSX.stat().st_mtime
api.refresh()
ok("更新報價不會改到 持股.xlsx", C.HOLDINGS_XLSX.stat().st_mtime == before)


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 78)
print("5. 手動匯率覆寫")
print("=" * 78)
make_holdings([["AAPL", "", "US", 10, 100, ""]],
              settings={**C.DEFAULT_SETTINGS, C.SET_MANUAL_FX: 33.5})
p2 = app_module.Api().refresh()
if p2.get("error"):
    print("  ERROR:", p2["error"]); FAIL += 1
else:
    ok("匯率用手動指定的值", p2["summary"]["fx_rate"] == 33.5, str(p2["summary"]["fx_rate"]))
    ok("來源標示為手動指定", p2["summary"]["fx_source"] == "手動指定")
    r = p2["rows"][0]
    ok("市值用 33.5 換算", abs(r["市值"] - 10 * r["現價"] * 33.5) < 0.01)


# ══════════════════════════════════════════════════════════════════
print("\n" + "=" * 78)
print("6. 上櫃股票（測 tse_ 失敗自動改 otc_）")
print("=" * 78)
make_holdings([["6488", "", "TW", 100, 500, ""], ["5483", "", "TW", 100, 100, ""]])
p3 = app_module.Api().refresh()
if p3.get("error"):
    print("  ERROR:", p3["error"]); FAIL += 1
else:
    for r in p3["rows"]:
        print(f"  {r['代號']}  {r['名稱'] or '?':<12} 現價 {r['現價']}  來源 {r['資料來源']}")
    ok("上櫃股票也抓得到", all(r["現價"] is not None for r in p3["rows"]))


shutil.rmtree(TMP, ignore_errors=True)
print("\n" + "=" * 78)
print("整合測試" + ("全部通過" if not FAIL else f"有 {FAIL} 項失敗"))
print("=" * 78)
sys.exit(1 if FAIL else 0)
