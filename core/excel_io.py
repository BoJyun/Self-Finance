"""Excel 讀寫層。

設計原則 —— 讀寫分離：
  * 持股.xlsx  由使用者擁有，程式「只讀不寫」（首次執行產生範本時例外）。
    這樣使用者可以一直開著 Excel 編輯，程式照樣能更新報價，不會撞到檔案鎖定。
  * data\\ 底下的檔案由程式擁有，使用者平常不會去開，所以程式寫得進去。

寫入一律採「先寫暫存檔、成功後才取代」的原子替換，中途當機不會留下壞掉的檔案。
"""
from __future__ import annotations

import datetime as _dt
import json
import os
import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config as C


class HoldingsFileError(Exception):
    """持股.xlsx 有問題，訊息會直接顯示給使用者，所以要寫得看得懂。"""


# ══════════════════════════════════════════════════════════════════
#  資料結構
# ══════════════════════════════════════════════════════════════════
@dataclass
class Holding:
    code: str
    name: str
    market: str          # C.MARKET_TW / C.MARKET_US
    shares: float
    avg_cost: float
    note: str = ""
    row: int = 0         # 在 Excel 裡的列號，出錯時好指給使用者看


@dataclass
class HoldingsFile:
    holdings: list[Holding] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    mtime: float = 0.0           # 讀取當下的檔案時間，寫回前用來偵測外部改動


# ══════════════════════════════════════════════════════════════════
#  小工具
# ══════════════════════════════════════════════════════════════════
def _to_number(value, *, field_name: str, row: int):
    """把儲存格轉成數字。容忍 "38,000"、" 30.57 "、"$180.5" 這類人工輸入。"""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "").replace("　", "")
    try:
        return float(text)
    except ValueError:
        raise HoldingsFileError(
            f"第 {row} 列的「{field_name}」是「{value}」，不是有效的數字。\n"
            f"請把它改成純數字（例如 38000 或 30.57）。"
        )


def infer_market(code: str) -> str:
    """代號純數字視為台股，含英文字母視為美股。"""
    return C.MARKET_TW if re.fullmatch(r"\d{4,6}[A-Z]?", code or "") else C.MARKET_US


def _norm_market(raw, code: str, row: int) -> str:
    if raw is None or not str(raw).strip():
        return infer_market(code)
    text = str(raw).strip().upper()
    if text in ("TW", "台股", "台", "TWSE", "TPEX"):
        return C.MARKET_TW
    if text in ("US", "美股", "美", "USA"):
        return C.MARKET_US
    raise HoldingsFileError(
        f"第 {row} 列的「{C.COL_MARKET}」是「{raw}」，看不懂。\n"
        f"請填 TW（台股）或 US（美股），或整格留空讓程式自動判斷。"
    )


# ══════════════════════════════════════════════════════════════════
#  讀取 持股.xlsx
# ══════════════════════════════════════════════════════════════════
def read_holdings(path: Path | None = None) -> HoldingsFile:
    path = path or C.HOLDINGS_XLSX
    if not path.exists():
        raise HoldingsFileError(
            f"找不到持股檔案：\n{path}\n\n"
            f"請關掉程式後重新啟動，程式會自動幫你建立一個範本檔。"
        )

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except PermissionError:
        raise HoldingsFileError(
            f"讀不到 {path.name}，檔案可能正被其他程式獨佔。\n"
            f"請確認沒有其他程式鎖住它，然後再試一次。"
        )
    except Exception as e:
        raise HoldingsFileError(
            f"開啟 {path.name} 失敗：{e}\n\n"
            f"檔案可能損毀。你可以從 {C.BACKUP_DIR} 找回之前的版本，"
            f"或是把它刪掉讓程式重新產生範本。"
        )

    try:
        return HoldingsFile(holdings=_parse_holdings_sheet(wb, path),
                            mtime=path.stat().st_mtime)
    finally:
        wb.close()


def _parse_holdings_sheet(wb, path: Path) -> list[Holding]:
    if C.SHEET_HOLDINGS not in wb.sheetnames:
        raise HoldingsFileError(
            f"{path.name} 裡找不到名為「{C.SHEET_HOLDINGS}」的分頁。\n"
            f"目前有的分頁：{'、'.join(wb.sheetnames)}\n\n"
            f"請把放持股資料的分頁改名為「{C.SHEET_HOLDINGS}」。"
        )
    ws = wb[C.SHEET_HOLDINGS]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HoldingsFileError(f"「{C.SHEET_HOLDINGS}」分頁是空的，第 1 列需要有欄位標題。")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header) if name}

    missing = [c for c in C.REQUIRED_COLUMNS if c not in idx]
    if missing:
        raise HoldingsFileError(
            f"「{C.SHEET_HOLDINGS}」分頁第 1 列缺少必要欄位：{'、'.join(missing)}\n\n"
            f"目前讀到的標題是：{'、'.join(h for h in header if h) or '(空白)'}\n"
            f"必要欄位：{'、'.join(C.REQUIRED_COLUMNS)}\n"
            f"選填欄位：{C.COL_NAME}、{C.COL_MARKET}、{C.COL_NOTE}"
        )

    def cell(row, col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    holdings: list[Holding] = []
    seen: dict[str, int] = {}
    for n, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue                                        # 跳過空白列

        code = str(cell(row, C.COL_CODE) or "").strip()
        if not code:
            continue                                        # 沒代號就當這列不存在
        if code.endswith(".0"):                             # Excel 把 2882 存成數字的情況
            code = code[:-2]

        shares = _to_number(cell(row, C.COL_SHARES), field_name=C.COL_SHARES, row=n)
        avg = _to_number(cell(row, C.COL_AVGCOST), field_name=C.COL_AVGCOST, row=n)
        if shares is None or avg is None:
            raise HoldingsFileError(
                f"第 {n} 列（{code}）的「{C.COL_SHARES}」或「{C.COL_AVGCOST}」是空的。\n"
                f"這兩欄都要填，不然算不出損益。整列不要的話請整列刪掉。"
            )
        if shares <= 0:
            raise HoldingsFileError(f"第 {n} 列（{code}）的股數是 {shares:g}，必須大於 0。")
        if avg < 0:
            raise HoldingsFileError(f"第 {n} 列（{code}）的均價是 {avg:g}，不能是負數。")

        if code in seen:
            raise HoldingsFileError(
                f"代號「{code}」出現了兩次（第 {seen[code]} 列和第 {n} 列）。\n"
                f"請把它們合併成一列，股數相加、均價填加權平均。"
            )
        seen[code] = n

        holdings.append(Holding(
            code=code,
            name=str(cell(row, C.COL_NAME) or "").strip(),
            market=_norm_market(cell(row, C.COL_MARKET), code, n),
            shares=shares,
            avg_cost=avg,
            note=str(cell(row, C.COL_NOTE) or "").strip(),
            row=n,
        ))
    return holdings


# ══════════════════════════════════════════════════════════════════
#  首次執行：產生範本
# ══════════════════════════════════════════════════════════════════
_HEADER_FILL = PatternFill("solid", fgColor="1F2430")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_FORMAT_SPARE_ROWS = 20        # 預先套用文字格式的空白列數


def _style_holdings_sheet(ws) -> None:
    for i, _ in enumerate(C.HOLDING_COLUMNS, start=1):
        ws.cell(row=1, column=i).fill = _HEADER_FILL
        ws.cell(row=1, column=i).font = _HEADER_FONT
    for i, w in enumerate([12, 22, 8, 12, 12, 44], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    # 代號欄設成文字格式，否則 Excel 會把 00919 變成 919。
    # 多留幾列空白也套用，這樣你直接在 Excel 手動加一列時也不會被吃掉開頭的 0。
    # 緩衝別開太大 —— 套過格式的儲存格會被算進 max_row，檔案會顯得有一堆空列。
    for r in range(2, ws.max_row + 1 + _FORMAT_SPARE_ROWS):
        ws.cell(row=r, column=1).number_format = "@"


def ensure_holdings_template(path: Path | None = None) -> bool:
    """持股.xlsx 不存在時建立範本。回傳 True 表示這次有建立。"""
    path = path or C.HOLDINGS_XLSX
    if path.exists():
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = C.SHEET_HOLDINGS
    ws.append(C.HOLDING_COLUMNS)
    ws.append(["2882", "", "TW", 217, 44.66, "範例：可直接改成你自己的持股"])
    ws.append(["AAPL", "", "US", 50, 180.5, "範例：美股用英文代號"])
    _style_holdings_sheet(ws)

    atomic_save(wb, path)
    return True


# ══════════════════════════════════════════════════════════════════
#  寫入
# ══════════════════════════════════════════════════════════════════
def atomic_save(wb: Workbook, path: Path) -> None:
    """先寫暫存檔，成功後才取代目標檔，避免寫到一半當機留下壞檔。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    os.replace(tmp, path)


def backup(path: Path) -> None:
    if not path.exists():
        return
    C.BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(path, C.BACKUP_DIR / f"{path.stem}_{stamp}{path.suffix}")

    olds = sorted(C.BACKUP_DIR.glob(f"{path.stem}_*{path.suffix}"))
    for old in olds[:-C.KEEP_BACKUPS]:
        try:
            old.unlink()
        except OSError:
            pass


# 舊名稱，內部還有幾處在用
_atomic_save = atomic_save
_backup = backup


def excel_lock_holder(path: Path) -> str | None:
    """Excel 開著某個檔案時會在同目錄留下 ~$檔名 的暫存檔。
    回傳該檔路徑字串代表「檔案疑似正被 Excel 開啟」，None 代表沒有。

    這只是提前警告 —— 真正的判斷還是靠寫入時的 PermissionError。
    """
    lock = path.with_name("~$" + path.name)
    return str(lock) if lock.exists() else None


def write_holdings(holdings: list[dict], path: Path | None = None,
                   expect_mtime: float | None = None) -> dict:
    """把編輯後的持股寫回 持股.xlsx。

    只動「持股」分頁的內容，其他分頁與欄寬等格式都保留。
    寫入前先備份，並用原子替換，中途失敗不會留下壞檔。
    """
    path = path or C.HOLDINGS_XLSX

    if expect_mtime is not None and path.exists():
        if abs(path.stat().st_mtime - expect_mtime) > 1.0:
            raise HoldingsFileError(
                f"{path.name} 在你編輯的期間被其他程式改過了。\n\n"
                f"為了避免蓋掉那些改動，這次沒有存檔。\n"
                f"請按「取消」離開編輯模式、重新讀取最新資料後再改一次。"
            )

    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = C.SHEET_HOLDINGS
    else:
        try:
            wb = load_workbook(path)
        except PermissionError:
            raise HoldingsFileError(_locked_message(path))
        except Exception as e:
            raise HoldingsFileError(f"開啟 {path.name} 失敗：{e}")
        ws = (wb[C.SHEET_HOLDINGS] if C.SHEET_HOLDINGS in wb.sheetnames
              else wb.create_sheet(C.SHEET_HOLDINGS, 0))

    # 清掉舊內容（保留欄寬、凍結窗格這些 sheet 層級的設定）
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.append(C.HOLDING_COLUMNS)
    for h in holdings:
        ws.append([
            str(h.get(C.COL_CODE, "")).strip(),
            str(h.get(C.COL_NAME, "") or "").strip(),
            str(h.get(C.COL_MARKET, "") or "").strip(),
            h.get(C.COL_SHARES),
            h.get(C.COL_AVGCOST),
            str(h.get(C.COL_NOTE, "") or "").strip(),
        ])
    _style_holdings_sheet(ws)

    try:
        backup(path)
        atomic_save(wb, path)
    except PermissionError:
        raise HoldingsFileError(_locked_message(path))
    finally:
        wb.close()

    return {"count": len(holdings), "path": str(path),
            "mtime": path.stat().st_mtime}


def _locked_message(path: Path) -> str:
    return (f"存不進 {path.name} —— 這個檔案正被 Excel 開著。\n\n"
            f"請先在 Excel 裡關掉它，再按一次儲存。\n"
            f"（你剛才改的內容還在，不會不見。）")


def append_nav_record(record: dict, path: Path | None = None) -> None:
    """每日淨值追加一筆；同一天重複執行就覆蓋當天那筆。"""
    path = path or C.HISTORY_XLSX
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        _backup(path)
        wb = load_workbook(path)
        ws = wb[C.SHEET_NAV] if C.SHEET_NAV in wb.sheetnames else wb.create_sheet(C.SHEET_NAV)
        if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
            ws.append(C.NAV_COLUMNS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = C.SHEET_NAV
        ws.append(C.NAV_COLUMNS)
        for i, _ in enumerate(C.NAV_COLUMNS, start=1):
            ws.cell(row=1, column=i).fill = _HEADER_FILL
            ws.cell(row=1, column=i).font = _HEADER_FONT
        for i, w in enumerate([12, 16, 16, 16, 12, 16, 18, 12], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    values = [record.get(c) for c in C.NAV_COLUMNS]
    today = str(record.get("日期"))

    target = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value) == today:
            target = r
            break
    if target is None:
        ws.append(values)
    else:
        for i, v in enumerate(values, start=1):
            ws.cell(row=target, column=i, value=v)

    _atomic_save(wb, path)


def export_overview(rows: list[dict], summary: dict, path: Path | None = None) -> Path:
    """把目前的庫存總覽匯出成 Excel（給使用者存檔／報稅／自己再加工用）。"""
    path = path or C.EXPORT_XLSX
    if path.exists():
        _backup(path)

    cols = ["代號", "名稱", "市場", "股數", "均價", "現價", "昨收", "漲跌幅",
            "今日損益", "總損益", "報酬率", "市值", "成本", "幣別", "資料來源", "備註"]

    wb = Workbook()
    ws = wb.active
    ws.title = "庫存總覽"
    ws.append([f"產生時間 {_dt.datetime.now():%Y-%m-%d %H:%M:%S}"
               f"    USD/TWD {summary.get('fx_rate')}"
               f"    {summary.get('fx_note', '')}"])
    ws.append([])
    ws.append(cols)
    header_row = 3
    for i, _ in enumerate(cols, start=1):
        ws.cell(row=header_row, column=i).fill = _HEADER_FILL
        ws.cell(row=header_row, column=i).font = _HEADER_FONT

    for row in rows:
        ws.append([row.get(c) for c in cols])

    ws.append([])
    ws.append(["總計", "", "", "", "", "", "", "",
               summary.get("today_pnl"), summary.get("total_pnl"),
               summary.get("total_return"), summary.get("market_value"),
               summary.get("cost"), "TWD", "", ""])

    for i, w in enumerate([10, 22, 8, 12, 12, 12, 12, 10, 14, 14, 10, 16, 16, 8, 14, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    _atomic_save(wb, path)
    return path


# ══════════════════════════════════════════════════════════════════
#  報價快取（開程式時先顯示上次的數字，不用空等網路）
# ══════════════════════════════════════════════════════════════════
def save_cache(payload: dict) -> None:
    C.DATA_DIR.mkdir(parents=True, exist_ok=True)
    tmp = C.CACHE_JSON.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    os.replace(tmp, C.CACHE_JSON)


def load_cache() -> dict | None:
    try:
        return json.loads(C.CACHE_JSON.read_text(encoding="utf-8"))
    except Exception:
        return None
