"""程式設定，存在 data\\設定.json。

刻意跟 持股.xlsx 分開：Excel 開著的時候寫不進去，
不該因此連「換個漲跌顏色」都做不到。
"""
from __future__ import annotations

import json
import os

from . import config as C


def _clean(raw: dict) -> dict:
    """把任何來源的值都正規化成可用的型別，壞值退回預設。"""
    s = dict(C.DEFAULT_SETTINGS)

    fx = raw.get(C.SET_MANUAL_FX)
    if fx is None or str(fx).strip() == "":
        s[C.SET_MANUAL_FX] = None
    else:
        try:
            v = float(str(fx).strip())
            s[C.SET_MANUAL_FX] = v if v > 0 else None
        except ValueError:
            s[C.SET_MANUAL_FX] = None

    s[C.SET_COLOR_SCHEME] = (C.COLOR_US_STYLE
                             if str(raw.get(C.SET_COLOR_SCHEME, "")).strip() == C.COLOR_US_STYLE
                             else C.COLOR_TW_STYLE)

    auto = raw.get(C.SET_AUTO_REFRESH, True)
    if isinstance(auto, str):
        auto = auto.strip().lower() not in ("", "0", "false", "off", "否", "關", "關閉")
    s[C.SET_AUTO_REFRESH] = bool(auto)
    return s


def load() -> dict:
    try:
        return _clean(json.loads(C.SETTINGS_JSON.read_text(encoding="utf-8")))
    except FileNotFoundError:
        return dict(C.DEFAULT_SETTINGS)
    except Exception:
        return dict(C.DEFAULT_SETTINGS)      # 檔案壞掉就用預設，不要讓程式開不起來


def save(raw: dict) -> dict:
    """寫回設定，回傳正規化後的結果。"""
    s = _clean(raw)
    C.DATA_DIR.mkdir(parents=True, exist_ok=True)
    tmp = C.SETTINGS_JSON.with_suffix(".tmp")
    tmp.write_text(json.dumps(s, ensure_ascii=False, indent=1), encoding="utf-8")
    os.replace(tmp, C.SETTINGS_JSON)
    return s


def migrate_from_excel() -> bool:
    """舊版把設定放在 持股.xlsx 的「設定」分頁。
    第一次執行時搬到 設定.json，然後把該分頁從 Excel 移除。
    回傳 True 表示這次有搬。
    """
    if C.SETTINGS_JSON.exists() or not C.HOLDINGS_XLSX.exists():
        return False

    from openpyxl import load_workbook

    try:
        wb = load_workbook(C.HOLDINGS_XLSX)
    except Exception:
        return False

    try:
        if C.SHEET_SETTINGS_LEGACY not in wb.sheetnames:
            return False

        raw: dict = {}
        for row in wb[C.SHEET_SETTINGS_LEGACY].iter_rows(values_only=True):
            if row and row[0] is not None and len(row) > 1:
                raw[str(row[0]).strip()] = row[1]

        # 舊的「自動更新秒數」是數字，0 代表關閉
        if C.LEGACY_AUTO_REFRESH in raw:
            try:
                raw[C.SET_AUTO_REFRESH] = float(raw[C.LEGACY_AUTO_REFRESH] or 0) > 0
            except (TypeError, ValueError):
                raw[C.SET_AUTO_REFRESH] = True

        save(raw)

        del wb[C.SHEET_SETTINGS_LEGACY]
        from .excel_io import atomic_save, backup
        backup(C.HOLDINGS_XLSX)
        atomic_save(wb, C.HOLDINGS_XLSX)
        return True
    except Exception:
        # 搬不動就算了，至少把設定檔建起來，下次不會再試
        if not C.SETTINGS_JSON.exists():
            save({})
        return False
    finally:
        wb.close()
